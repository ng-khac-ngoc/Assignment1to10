import bs4, requests, datetime, re, csv
import openpyxl

dt = datetime.datetime.now()
url = 'https://www.cisa.gov/uscert/ncas/alerts'

csvFile = open('Assign5.csv', 'w', newline='')              # open file
csvWriter = csv.writer(csvFile, delimiter=',', lineterminator='\n')            # write file
csvWriter.writerow(['Alert ID', 'Alert Name', 'Release Date', 'Last revised', 'Tips', 'Alert Link'])        # add the first row


def addDataToCsv(csvFile, title, subTitle, releaseDate1, lastRevised, alertLink):      # add data in file
    csvFile.writerow([title, subTitle, releaseDate1, lastRevised, ' ', alertLink])


wb = openpyxl.Workbook()                # add title in the A1 -> F1
sheet = wb.active
sheet['A1'] = 'Alert ID'
sheet['B1'] = 'Alert Name'
sheet['C1'] = 'Release Date'
sheet['D1'] = 'Last revised'
sheet['E1'] = 'Tips'
sheet['F1'] = 'Alert Link'
sheet.column_dimensions['A'].width = 10
sheet.column_dimensions['B'].width = 125
sheet.column_dimensions['C'].width = 20
sheet.column_dimensions['D'].width = 20
sheet.column_dimensions['E'].width = 10
sheet.column_dimensions['F'].width = 30


def addDataToExcel(curentRow, title, subTitle, releaseDate1, lastRevised, alertLink):       # add data to excel
    sheet.cell(row=curentRow, column=1, value=title)
    sheet.cell(row=curentRow, column=2, value=subTitle)
    sheet.cell(row=curentRow, column=3, value=releaseDate1)
    sheet.cell(row=curentRow, column=4, value=lastRevised)
    sheet.cell(row=curentRow, column=5, value='')
    sheet.cell(row=curentRow, column=6, value=alertLink)


def parserHtml(url):                        # get url and parser HTML
    try:
        res = requests.get(url)
        res.raise_for_status()
    except:
        print('url error')
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
    return soup


elem = parserHtml(url).select('.date-listing a')[0].get('href')
year = parserHtml(url).select('.date-listing a')[0].text                # 2022

# soup3 = parserHtml(url).select('.tips-link h3 a')[0].get('href')

curentRow = 2

if str(dt.year) == year:
    url2 = 'https://www.cisa.gov/uscert' + elem
    soup2 = parserHtml(url2).select('.field-content a')
    # soup3 = parserHtml(url2).select('.tips-link h3 a')[0].get('href')

    for s in soup2:
        url3 = 'https://www.cisa.gov/uscert' + s.get('href')
        title = re.search('\((.*)\)',(parserHtml(url3).select('#page-title')[0].text)).group(1)     # get page title

        subTitle = parserHtml(url3).select('#page-sub-title')[0].text           # get sub title

        releaseDate = parserHtml(url3).select('.submitted.meta-text')[0].contents               # get release date
        lenReleaseDate = len(releaseDate)

        if lenReleaseDate == 1:
            releaseDate1 = re.search(r'Original release date: (.*)', (parserHtml(url3).select('.submitted.meta-text')[0].text.strip())).group(1)
            lastRevised = releaseDate1
        else:
            releaseDate1 = re.search(r'Original release date: ([^|]*)', (parserHtml(url3).select('.submitted.meta-text')[0].text.strip())).group(1)
            lastRevised = re.search(r'(.*)(Last revised: )(.*)', (parserHtml(url3).select('.submitted.meta-text')[0].text.strip())).group(3)

        alertLink = parserHtml(url3).select('link')[2].get('href')          # get alertLink

        addDataToCsv(csvWriter, title , subTitle, releaseDate1, lastRevised, alertLink)    # add data to csv file

        addDataToExcel(curentRow, title, subTitle, releaseDate1, lastRevised, alertLink)    # add data to Excel file

        curentRow += 1


csvFile.close()
wb.save('Assign5Excel.xlsx')

